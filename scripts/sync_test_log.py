"""
sync_test_log.py — Tarefa 2.7: Log de testes (planilha controle)

Sincroniza métricas do Neon DB com a planilha
"T7 - Registro de Testes de Criativos.xlsx".

Lógica:
- Preserva: Hipótese, Aprendizado, Veredicto manual, LPV, Connect/Conv Rate
- Atualiza: Spend, Impressões, CPM, CTR, Cliques, Leads, CPL por criativo
- Adiciona: novos criativos do DB que ainda não estão na planilha
- Aba Dados Brutos: substituída por métricas diárias do DB
- Aba Aprendizados: atualiza timestamp

Uso:
  pip install psycopg2-binary openpyxl --break-system-packages
  DATABASE_URL=<neon_url> python scripts/sync_test_log.py [--xlsx PATH] [--dry-run]

Variáveis de ambiente:
  DATABASE_URL   — Neon connection string (obrigatório)
  XLSX_PATH      — Caminho do arquivo xlsx (padrão: Documents--Pegasus Ads/T7 - Registro de Testes de Criativos.xlsx)
"""

import os
import sys
import argparse
from datetime import date
import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuração ──────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)

# Tentar localizar o xlsx em caminhos comuns
POSSIBLE_XLSX_PATHS = [
    os.path.join(REPO_ROOT, "..", "Documents--Pegasus Ads", "T7 - Registro de Testes de Criativos.xlsx"),
    os.path.join(REPO_ROOT, "..", "claude--Pegasus Ads", "T7 - Registro de Testes de Criativos.xlsx"),
    r"C:\Users\leand\Documents\Pegasus Ads\T7 - Registro de Testes de Criativos.xlsx",
]

# Colunas da aba Criativos (1-indexed)
COL_NOME       = 1   # A - Criativo (KEY)
COL_TIPO       = 2   # B - Tipo
COL_IA         = 3   # C - IA
COL_PARCERIA   = 4   # D - Parceria
COL_CAMPANHAS  = 5   # E - Campanhas
COL_ADSETS     = 6   # F - Ad Sets
COL_SPEND      = 7   # G - Spend Total  ← atualizar
COL_IMPRESSOES = 8   # H - Impressões   ← atualizar
COL_CPM        = 9   # I - CPM Médio    ← atualizar
COL_CTR        = 10  # J - CTR          ← atualizar
COL_CLIQUES    = 11  # K - Cliques      ← atualizar
COL_LPV        = 12  # L - LPV          ← preservar
COL_CONNECT    = 13  # M - Connect Rate ← preservar
COL_LEADS      = 14  # N - Leads        ← atualizar
COL_CPL        = 15  # O - CPL          ← atualizar
COL_CONV       = 16  # P - Conv Rate    ← preservar
COL_VEREDITO   = 17  # Q - Veredicto    ← atualizar se kill rule ativa E veredito atual vazio
COL_HIPOTESE   = 18  # R - Hipótese     ← preservar
COL_APRENDIZADO= 19  # S - Aprendizado  ← preservar

VEREDICTOS_MANUAIS = {"Vencedor", "Ruim", "Caro", "Aceitável", "Sem dados", "Em teste", "Em andamento"}
VEREDICTOS_AUTO    = {"Kill L0", "Kill L1", "Kill L2"}

CPL_TARGET = 25.0

# Cores para veredictos
CORES_VEREDITO = {
    "Vencedor":    "C6EFCE",  # verde
    "Kill L0":     "FFC7CE",  # vermelho
    "Kill L1":     "FFC7CE",  # vermelho
    "Kill L2":     "FFC7CE",  # vermelho
    "Ruim":        "FFC7CE",  # vermelho
    "Em teste":    "BDD7EE",  # azul
    "Caro":        "FFEB9C",  # amarelo
    "Aceitável":   "FFEB9C",  # amarelo
    "Sem dados":   "D9D9D9",  # cinza
    "Em andamento":"BDD7EE",  # azul
}


# ── Helper: normalizar nome do criativo ───────────────────────────────────────

def get_base_name(name: str) -> str:
    """Remove extensão e sufixo F/S do nome do criativo."""
    # Remove extensão
    if "." in name:
        name = name.rsplit(".", 1)[0]
    # Remove sufixo F ou S no final
    if name and name[-1] in ("F", "S"):
        name = name[:-1]
    return name


# ── Conexão DB ────────────────────────────────────────────────────────────────

def get_db(db_url: str):
    conn = psycopg2.connect(db_url)
    conn.autocommit = False
    return conn


def fetch_creative_metrics(conn) -> dict:
    """
    Retorna dicionário: base_name → métricas agregadas.
    """
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT
            c.name,
            c.status,
            c.generation,
            SUM(m.spend)        AS total_spend,
            SUM(m.impressions)  AS total_impressions,
            SUM(m.clicks)       AS total_clicks,
            SUM(m.leads)        AS total_leads,
            AVG(m.cpm)          AS avg_cpm,
            AVG(m.ctr)          AS avg_ctr,
            COUNT(DISTINCT m.date) AS days_count
        FROM creatives c
        LEFT JOIN metrics m ON m.creative_id = c.id
        GROUP BY c.id, c.name, c.status, c.generation
        ORDER BY c.generation ASC, c.created_at ASC
    """)
    result = {}
    for row in cur.fetchall():
        base = get_base_name(row["name"])
        spend = float(row["total_spend"] or 0)
        leads = int(row["total_leads"] or 0)
        impressions = int(row["total_impressions"] or 0)
        clicks = int(row["total_clicks"] or 0)
        cpl = round(spend / leads, 2) if leads > 0 else None
        result[base] = {
            "name": row["name"],
            "status": row["status"],
            "generation": row["generation"],
            "spend": round(spend, 2),
            "impressions": impressions,
            "clicks": clicks,
            "leads": leads,
            "cpm": round(float(row["avg_cpm"] or 0), 2),
            "ctr": round(float(row["avg_ctr"] or 0), 6),
            "cpl": cpl,
            "has_metrics": spend > 0,
        }
    cur.close()
    return result


def fetch_daily_metrics(conn) -> list:
    """
    Retorna lista de métricas diárias por criativo para a aba Dados Brutos.
    """
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT
            c.name,
            m.date,
            m.spend,
            m.impressions,
            m.cpm,
            m.ctr,
            m.clicks,
            m.leads,
            m.cpl,
            m.meta_ad_id
        FROM creatives c
        JOIN metrics m ON m.creative_id = c.id
        ORDER BY c.name ASC, m.date ASC
    """)
    rows = cur.fetchall()
    cur.close()
    return [dict(r) for r in rows]


def evaluate_kill_rule(metrics: dict) -> str | None:
    """Avalia L0-L2 e retorna string do veredito ou None."""
    spend = metrics.get("spend", 0)
    leads = metrics.get("leads", 0)
    cpl = metrics.get("cpl")

    if leads == 0 and spend >= CPL_TARGET * 1.5:
        return "Kill L0"
    if leads > 0 and cpl and cpl > CPL_TARGET * 3:
        return "Kill L1"
    if leads > 0 and cpl and cpl > CPL_TARGET * 2:
        return "Kill L2"
    return None


# ── Formatar célula de métrica ────────────────────────────────────────────────

def fmt_currency(ws, row, col, value):
    """Escreve valor monetário (R$)."""
    if value is None or value == 0:
        ws.cell(row=row, column=col).value = None
    else:
        c = ws.cell(row=row, column=col)
        c.value = value
        c.number_format = '#,##0.00'


def fmt_pct(ws, row, col, value):
    """Escreve percentual."""
    if value is None or value == 0:
        ws.cell(row=row, column=col).value = None
    else:
        c = ws.cell(row=row, column=col)
        c.value = value
        c.number_format = '0.00%'


def fmt_int(ws, row, col, value):
    """Escreve inteiro."""
    ws.cell(row=row, column=col).value = value if value else None


def apply_veredito_color(cell, veredito: str):
    cor = CORES_VEREDITO.get(veredito)
    if cor:
        cell.fill = PatternFill("solid", start_color=cor, fgColor=cor)


# ── Atualizar aba Criativos ────────────────────────────────────────────────────

def sync_criativos(ws, db_metrics: dict, dry_run: bool) -> dict:
    stats = {"updated": 0, "added": 0, "unchanged": 0, "kill_applied": 0}

    # Mapear linhas existentes: base_name → row_index
    existing = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[COL_NOME - 1]
        if cell.value:
            base = get_base_name(str(cell.value))
            existing[base] = cell.row

    # 1. Atualizar criativos que já existem
    for base_name, row_idx in existing.items():
        if base_name not in db_metrics:
            stats["unchanged"] += 1
            continue

        m = db_metrics[base_name]
        if not m["has_metrics"]:
            stats["unchanged"] += 1
            continue

        if not dry_run:
            fmt_currency(ws, row_idx, COL_SPEND, m["spend"])
            fmt_int(ws, row_idx, COL_IMPRESSOES, m["impressions"])
            fmt_currency(ws, row_idx, COL_CPM, m["cpm"])
            fmt_pct(ws, row_idx, COL_CTR, m["ctr"])
            fmt_int(ws, row_idx, COL_CLIQUES, m["clicks"])
            fmt_int(ws, row_idx, COL_LEADS, m["leads"])
            fmt_currency(ws, row_idx, COL_CPL, m["cpl"])

            # Aplicar kill rule automaticamente se veredito está vazio
            current_veredito = ws.cell(row_idx, COL_VEREDITO).value or ""
            kill = evaluate_kill_rule(m)
            if kill and (not current_veredito or current_veredito in VEREDICTOS_AUTO):
                veredito_cell = ws.cell(row_idx, COL_VEREDITO)
                veredito_cell.value = kill
                apply_veredito_color(veredito_cell, kill)
                stats["kill_applied"] += 1

        stats["updated"] += 1

    # 2. Adicionar novos criativos do DB
    next_row = ws.max_row + 1
    for base_name, m in db_metrics.items():
        if base_name in existing:
            continue

        # Criativo novo — adicionar linha
        if not dry_run:
            ws.cell(next_row, COL_NOME).value = base_name
            ws.cell(next_row, COL_TIPO).value = "Imagem"

            if m["has_metrics"]:
                fmt_currency(ws, next_row, COL_SPEND, m["spend"])
                fmt_int(ws, next_row, COL_IMPRESSOES, m["impressions"])
                fmt_currency(ws, next_row, COL_CPM, m["cpm"])
                fmt_pct(ws, next_row, COL_CTR, m["ctr"])
                fmt_int(ws, next_row, COL_CLIQUES, m["clicks"])
                fmt_int(ws, next_row, COL_LEADS, m["leads"])
                fmt_currency(ws, next_row, COL_CPL, m["cpl"])

                kill = evaluate_kill_rule(m)
                veredito = kill or "Em teste"
                veredito_cell = ws.cell(next_row, COL_VEREDITO)
                veredito_cell.value = veredito
                apply_veredito_color(veredito_cell, veredito)
            else:
                ws.cell(next_row, COL_VEREDITO).value = "Em teste"

            ws.cell(next_row, COL_HIPOTESE).value = "— (novo)"
            ws.cell(next_row, COL_APRENDIZADO).value = "—"

        stats["added"] += 1
        next_row += 1

    return stats


# ── Atualizar aba Dados Brutos ────────────────────────────────────────────────

def sync_dados_brutos(ws, daily_metrics: list, dry_run: bool) -> int:
    if dry_run:
        return len(daily_metrics)

    # Limpar dados existentes (manter header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Reescrever com dados do DB
    row_idx = 2
    for m in daily_metrics:
        base = get_base_name(m["name"])
        leads = m["leads"] or 0
        spend = float(m["spend"] or 0)
        cpl = round(spend / leads, 3) if leads > 0 else None

        ws.cell(row_idx, 1).value = "T7__0003"   # campanha padrão
        ws.cell(row_idx, 2).value = base
        ws.cell(row_idx, 3).value = m.get("meta_ad_id", "")
        ws.cell(row_idx, 4).value = ""            # Status Ad (não disponível via DB)
        ws.cell(row_idx, 5).value = ""            # Status AdSet
        ws.cell(row_idx, 6).value = round(spend, 2)
        ws.cell(row_idx, 7).value = m["impressions"] or 0
        ws.cell(row_idx, 8).value = round(float(m["cpm"] or 0), 4)
        ws.cell(row_idx, 9).value = round(float(m["ctr"] or 0), 8)
        ws.cell(row_idx, 10).value = m["clicks"] or 0
        ws.cell(row_idx, 11).value = ""           # LPV — não temos no DB
        ws.cell(row_idx, 12).value = leads
        ws.cell(row_idx, 13).value = cpl
        ws.cell(row_idx, 14).value = ""           # Connect Rate
        ws.cell(row_idx, 15).value = ""           # Conv Rate
        ws.cell(row_idx, 16).value = "Não"        # Vídeo
        ws.cell(row_idx, 17).value = "Não"        # IA
        ws.cell(row_idx, 18).value = "Não"        # Parceria
        row_idx += 1

    return len(daily_metrics)


# ── Atualizar timestamp nos Aprendizados ──────────────────────────────────────

def update_timestamp(ws_aprendizados, dry_run: bool):
    today_str = date.today().strftime("%d/%m/%Y")
    for row in ws_aprendizados.iter_rows(min_row=1, max_row=5):
        for cell in row:
            val = str(cell.value or "")
            if "Última atualização" in val or "CPL Meta" in val:
                if not dry_run:
                    cell.value = f"CPL Meta: R${CPL_TARGET:.2f} | Última atualização: {today_str}"
                return


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Sincronizar métricas do DB com planilha de testes")
    parser.add_argument("--xlsx", help="Caminho do arquivo xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Apenas exibir o que seria feito sem salvar")
    args = parser.parse_args()

    db_url = os.environ.get("DATABASE_URL") or os.environ.get("POSTGRES_URL")
    if not db_url:
        print("❌ DATABASE_URL ou POSTGRES_URL não definido.")
        sys.exit(1)

    # Localizar xlsx
    xlsx_path = args.xlsx
    if not xlsx_path:
        for p in POSSIBLE_XLSX_PATHS:
            if os.path.exists(p):
                xlsx_path = p
                break
    if not xlsx_path or not os.path.exists(xlsx_path):
        # Tentar no workspace mnt
        fallback = os.path.join(REPO_ROOT, "..", "Documents--Pegasus Ads", "T7 - Registro de Testes de Criativos.xlsx")
        if not os.path.exists(fallback):
            print(f"❌ Arquivo xlsx não encontrado. Use --xlsx <caminho>")
            sys.exit(1)
        xlsx_path = fallback

    print(f"📊 Planilha: {xlsx_path}")
    print(f"🗄️  DB: {db_url[:40]}...")
    if args.dry_run:
        print("🔍 DRY RUN — nenhuma alteração será salva\n")

    # Conectar ao DB
    conn = get_db(db_url)
    print("✅ DB conectado")

    db_metrics = fetch_creative_metrics(conn)
    daily_metrics = fetch_daily_metrics(conn)
    conn.close()

    print(f"📦 Criativos no DB: {len(db_metrics)}")
    print(f"📅 Registros diários: {len(daily_metrics)}\n")

    # Carregar xlsx
    wb = load_workbook(xlsx_path)
    ws_criativos = wb["Criativos"]
    ws_brutos = wb["Dados Brutos"]
    ws_aprendizados = wb["Aprendizados"]

    # Sincronizar
    stats = sync_criativos(ws_criativos, db_metrics, args.dry_run)
    brutos_count = sync_dados_brutos(ws_brutos, daily_metrics, args.dry_run)
    update_timestamp(ws_aprendizados, args.dry_run)

    print(f"Criativos → Atualizados: {stats['updated']}  Adicionados: {stats['added']}  Kill aplicado: {stats['kill_applied']}  Sem mudança: {stats['unchanged']}")
    print(f"Dados Brutos → {brutos_count} linhas escritas")

    if not args.dry_run:
        wb.save(xlsx_path)
        print(f"\n✅ Planilha salva em: {xlsx_path}")
    else:
        print("\n(dry-run: nenhuma alteração salva)")


if __name__ == "__main__":
    main()
